import os
import sys
import re
import struct
import json
import logging
from pathlib import Path
from openpyxl import load_workbook
from jinja2 import Template

# 配置彩色日志
import colorama
from colorama import Fore, Style
colorama.init()

class ColoredFormatter(logging.Formatter):
    """带颜色输出的日志格式化器"""
    def format(self, record):
        color = Style.RESET_ALL
        if record.levelno >= logging.ERROR:
            color = Fore.RED
        elif record.levelno >= logging.WARNING:
            color = Fore.YELLOW
        
        message = super().format(record)
        return f"{color}{message}{Style.RESET_ALL}"

def setup_logging():
    """初始化日志配置"""
    handler = logging.StreamHandler()
    handler.setFormatter(
        ColoredFormatter(
            fmt='%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
    )
    
    logging.basicConfig(
        level=logging.INFO,
        handlers=[handler]
    )

setup_logging()

class TypeConverter:
    TYPE_MAPPING = {
        'int': {'ue': 'int32', 'fmt': '<i', 'size': 4},
        'int32': {'ue': 'int32', 'fmt': '<i', 'size': 4},
        'int64': {'ue': 'int64', 'fmt': '<q', 'size': 8},
        'float': {'ue': 'float', 'fmt': '<f', 'size': 4},
        'string': {'ue': 'FString', 'fmt': None, 'size': None},
        'bool': {'ue': 'bool', 'fmt': '?', 'size': 1}
    }

    @classmethod
    def parse_type(cls, type_str):
        """解析字段类型"""
        # 标准化类型输入
        type_str = type_str.strip().lower()
        if type_str == 'int': type_str = 'int32'
        elif type_str == 'str': type_str = 'string'
        
       # 支持 array<int32> 格式
        array_match = re.match(r'^array<(\w+)>$', type_str)
        if array_match:
            base_type = array_match.group(1)
            return ('array', base_type)
        
        # 处理Map类型
        if match := re.match(r'map<([\w]+),([\w]+)>', type_str):
            return ('map', match.group(1), match.group(2))
        
        # 基础类型
        if type_str in cls.TYPE_MAPPING:
            return ('basic', type_str)
        
        raise ValueError(f"Unsupported type: {type_str}")

    @classmethod
    def get_ue_type(cls, type_info):
        """获取UE类型表示"""
        if type_info[0] == 'array':
            return f"TArray<{cls.TYPE_MAPPING[type_info[1]]['ue']}>"
        if type_info[0] == 'map':
            return f"TMap<{cls.TYPE_MAPPING[type_info[1]]['ue']}, {cls.TYPE_MAPPING[type_info[2]]['ue']}>"
        return cls.TYPE_MAPPING[type_info[1]]['ue']
    
    @classmethod
    def _alias_type(cls, raw_type):
        """类型别名转换"""
        return {
            'int': 'int32',
            'str': 'string',
            'text': 'string',
            'boolean': 'bool'
        }.get(raw_type, raw_type)
    
    @classmethod
    def cast_value(cls, data, field_type):
        """增强的类型转换方法，处理空值"""
        # 处理空字符串和None值
        if not str(data).strip():
            if field_type in ['int32', 'int64']:
                return 0
            elif field_type == 'float':
                return 0.0
            elif field_type == 'bool':
                return False
            elif field_type == 'string':
                return ''
        
        try:
            if field_type == 'int32':
                return int(str(data).strip())
            if field_type == 'int64':
                return int(str(data).strip())
            if field_type == 'float':
                return float(str(data).strip())
            if field_type == 'bool':
                data_str = str(data).lower().strip()
                return data_str in ['true', '1', 'yes', 'y']
            return data
        except ValueError as e:
            raise ValueError(f"无法转换值 '{data}' 到类型 {field_type}: {str(e)}")



def get_column_letter(column_idx):
        """将数字列号转换为Excel列字母标号"""
        if not isinstance(column_idx, int) or column_idx < 1:
            raise ValueError("列号必须是从1开始的整数")
    
        letters = []
        while column_idx > 0:
            column_idx, remainder = divmod(column_idx - 1, 26)
            letters.append(chr(65 + remainder))  # 65是'A'的ASCII码
        return ''.join(reversed(letters))

class BinarySerializer:
    @staticmethod
    def serialize(data, field_type):
        """序列化数据"""
        type_info = TypeConverter.parse_type(field_type)
        
        if type_info[0] == 'array':
            return BinarySerializer._serialize_array(data, type_info[1])
        if type_info[0] == 'map':
            return BinarySerializer._serialize_map(data, *type_info[1:])
        return BinarySerializer._serialize_basic(data, type_info[1])

    @staticmethod
    def _serialize_array(data, element_type):
        buffer = struct.pack('<I', len(data))
        for elem in data:
            buffer += BinarySerializer.serialize(elem, element_type)
        return buffer

    @staticmethod
    def _serialize_map(data, key_type, val_type):
        buffer = struct.pack('<I', len(data))
        for k, v in data.items():
            buffer += BinarySerializer.serialize(k, key_type)
            buffer += BinarySerializer.serialize(v, val_type)
        return buffer

    @staticmethod
    def _serialize_basic(data, field_type):
        type_info = TypeConverter.TYPE_MAPPING[field_type]
        
        if field_type == 'string':
            encoded = data.encode('utf-8')
            return struct.pack('<I', len(encoded)) + encoded
        if field_type == 'bool':
            # 统一bool类型处理
            val = TypeConverter.cast_value(data, field_type)
            return struct.pack('?', val)
        
        # 数值类型统一转换
        converted = TypeConverter.cast_value(data, field_type)
        return struct.pack(type_info['fmt'], converted)

class ConfigConverter:
    MAGIC_NUMBER =  0x42434647  # 'BCFG'的ASCII码十六进制表示(大端)
    VERSION = 3
    
    UE_CONFIG_MGR_HEADER = Template('''#pragma once
#include "CoreMinimal.h"
#include "Kismet/BlueprintFunctionLibrary.h"
{% for header in headers %}
#include "GenConfig/{{ header }}Conf.h"
{% endfor %}
#include "ConfigMgr.generated.h"
UCLASS()
class COMMONCONFIG_API UConfigMgr : public UBlueprintFunctionLibrary
{
    GENERATED_BODY()

public:
    UFUNCTION(BlueprintCallable, Category = "ConfigMgr", meta = (WorldContext = "WorldContextObject"))
    static bool InitAllConfig(const UObject* WorldContextObject, const FString& ConfigPath);

    {% for base in base_names %}
    UFUNCTION(BlueprintPure, Category = "ConfigMgr")
    static void Get{{ base }}Config(int32 ID, bool& Found, F{{ base }}Conf& OutConfig);
    
    UFUNCTION(BlueprintPure, Category = "ConfigMgr")
    static TArray<int32> Get{{ base }}ConfigIDs();
    {% endfor %}
};
''')

    UE_CONFIG_MGR_CPP = Template('''#include "GenConfig/ConfigMgr.h"
#include "HAL/FileManager.h"

bool UConfigMgr::InitAllConfig(const UObject* WorldContextObject, const FString& ConfigPath)
{
    bool bAllSuccess = true;
    
    {% for base in base_names %}
    if(!{{ base }}ConfManager::getMe().Load(ConfigPath)) 
    {
        UE_LOG(LogTemp, Error, TEXT("Failed to load config: %s/%s.data"), *ConfigPath, TEXT("{{ base }}"));
        bAllSuccess = false;
    }
    {% endfor %}
    
    if(bAllSuccess)
    {
        UE_LOG(LogTemp, Log, TEXT("All configs loaded successfully!"));
    }
    return bAllSuccess;
}

{% for base in base_names %}
void UConfigMgr::Get{{ base }}Config(int32 ID, bool& Found, F{{ base }}Conf& OutConfig)
{
    const F{{ base }}Conf* Config = {{ base }}ConfManager::getMe().GetConfig(ID);
    if(Config)
    {
        Found = true;
        OutConfig = *Config;
    }
    else
    {
        Found = false;
    }
}

TArray<int32> UConfigMgr::Get{{ base }}ConfigIDs()
{
    return {{ base }}ConfManager::getMe().GetAllIDs();
}
{% endfor %}
''')
    
    UE_MANAGER_TEMPLATE = Template('''#pragma once
#include "CoreMinimal.h"
#include "TSingleton.h"                                   
#include "Containers/Map.h"
#include "Containers/Array.h"
#include "{{ header_file }}Conf.generated.h"

USTRUCT(BlueprintType)
struct F{{ struct_name }}
{
    GENERATED_BODY()
    {% for field in fields %}
    /* {{ field.comment }} */
    UPROPERTY(EditAnywhere, BlueprintReadWrite)
    {{ field.ue_type }} {{ field.name }}{% if field.ue_type == 'FString' %} = TEXT(""){% elif field.ue_type == 'int32' %} = 0{% endif %};
    {% endfor %}
};

class  {{ manager_class }} : public TSingleton<{{ manager_class }}> 
{
public:
    {{ manager_class }}() = default;
    ~{{ manager_class }}() { Clear(); }

    bool Load(const FString& ConfigPath);
    void Clear();

    const F{{ struct_name }}* GetConfig(int32 ID) const;
    const TArray<F{{ struct_name }}*>& GetAllConfigs() const { return Configs; }
    TArray<int32> GetAllIDs() const { TArray<int32> IDs;ConfigMap.GenerateKeyArray(IDs);return IDs; }
private:
    static void ParseConfig(FArchive& Ar, F{{ struct_name }}& Config);

    TMap<int32, F{{ struct_name }}*> ConfigMap;
    TArray<F{{ struct_name }}*> Configs;
};
''')
    
    UE_MANAGER_CPP_TEMPLATE = Template('''#include "GenConfig/{{ header_file }}Conf.h"
#include "Misc/FileHelper.h"
#include <string>
#include "Serialization/MemoryReader.h"                                       

bool {{ manager_class }}::Load(const FString& ConfigPath)
{
    Clear();
    TArray<uint8> FileData;
    FString FinalPath = FPaths::Combine(ConfigPath, TEXT("{{ data_file }}.data"));
    if(!FFileHelper::LoadFileToArray(FileData, *FinalPath)) return false;

    FMemoryReader Ar(FileData);
    // 验证文件头
    int32 Magic = 0;
    Ar << Magic;
    if(Magic != 0x42434647) return false;  // BCFG magic
    
    int32 Version = 0;
    Ar << Version;
    if(Version != 3) return false;

    int32 RecordCount = 0;
    Ar << RecordCount;

    for(int32 i = 0; i < RecordCount; ++i)
    {
        // 读取记录长度头
        int32 RecordLength = 0;
        Ar << RecordLength;
        // 验证长度有效性
        if(RecordLength <= 0 || Ar.TotalSize() - Ar.Tell() < RecordLength)
        {
            UE_LOG(LogTemp, Error, TEXT("Invalid record length: %d"), RecordLength);
            return false;
        }                                                              
         // 提取记录数据块
        TArray<uint8> RecordBuffer;
        RecordBuffer.SetNumUninitialized(RecordLength);
        Ar.Serialize(RecordBuffer.GetData(), RecordLength);

        // 创建子阅读器
        FMemoryReader RecordAr(RecordBuffer);
    

        F{{ struct_name }}* Item = new F{{ struct_name }}();
        ParseConfig(RecordAr, *Item);
      
        Configs.Add(Item);
        // 自动映射主键字段（假设第一个字段是ID）
        if(i == 0 && Configs.Num() > 0) {
            ConfigMap.Add(Item->{{ fields[0].name }}, Item);
        }
    }

    return true;
}

void {{ manager_class }}::ParseConfig(FArchive& Ar, F{{ struct_name }}& Config)
{
    {% for field in fields -%}
    {% if field.ue_type == 'FString' -%}
    // FString特殊处理
    {
        int32 StringLength;
        Ar << StringLength;
        if(StringLength > 0) 
        {
            std::string UTF8Data;
            UTF8Data.resize(StringLength,0);                           
            Ar.Serialize(UTF8Data.data(), StringLength);
            Config.{{ field.name }} = UTF8_TO_TCHAR(UTF8Data.c_str());
        }
    }
    {% elif field.type_info[0] == 'array' -%}
    // Array: {{ field.name }}
    {
        int32 Count = 0;
        Ar << Count;
        for(int32 i=0; i<Count; ++i){
            {{ field.type_info[1]|upper }} Element;
            {% if field.type_info[1] == 'FString' -%}
            // 嵌套FString处理
            int32 StrLen;
            Ar << StrLen;
            std::string StrData;
            StrData.resize(StrLen,0);                           
            Ar.Serialize(StrData.data(), StrLen);
            Element = UTF8_TO_TCHAR(StrData.c_str());
            {% else -%}
            Ar << Element;
            {% endif -%}
            Config.{{ field.name }}.Add(Element);
        }
    }
    {% elif field.type_info[0] == 'map' -%}
    // Map: {{ field.name }}
    {
        int32 Count = 0;
        Ar << Count;
        for(int32 i=0; i<Count; ++i){
            {{ field.type_info[1]|upper }} Key;
            {{ field.type_info[2]|upper }} Value;
            {% if field.type_info[1] == 'FString' -%}
            // Key是FString
            int32 KeyLen;
            Ar << KeyLen;
            std::string KeyData;
            KeyData.resize(KeyLen,0);
            Ar.Serialize(KeyData.data(), KeyLen);
            Key = UTF8_TO_TCHAR(KeyData.c_str());
            {% else %}
            Ar << Key;
            {% endif %}
            {% if field.type_info[2] == 'FString' -%}
            // Value是FString
            int32 ValLen;
            Ar << ValLen;
            std::string  ValData;
            ValData.resize(ValLen,0);                           
            Ar.Serialize(ValData.data(), ValLen);
            Value = UTF8_TO_TCHAR(ValData.c_str());
            {% else %}
            Ar << Value;
            {% endif %}
            Config.{{ field.name }}.Add(Key, Value);
        }
    }
    {% else -%}
    Ar << Config.{{ field.name }};
    {% endif -%}
    {% endfor %}
}

void {{ manager_class }}::Clear()
{
    ConfigMap.Empty();
    for(auto* Config : Configs) delete Config;
    Configs.Empty();
}

const F{{ struct_name }}* {{ manager_class }}::GetConfig(int32 ID) const
{
    const auto Found = ConfigMap.Find(ID);
    return Found ? *Found : nullptr;
}
''')
    
    UE_TEMPLATE = Template('''#pragma once
#include "CoreMinimal.h"
#include "Containers/Map.h"
#include "Containers/Array.h"
#include "{{ struct_name }}Conf.generated.h"
USTRUCT(BlueprintType)
struct F{{ struct_name }}Conf
{
    GENERATED_BODY()
    {% for field in fields %}
    // {{ field.comment }}
    UPROPERTY(EditAnywhere, BlueprintReadWrite)
    {{ field.ue_type }} {{ field.name }};
    {% endfor %}
};''')

    LUA_TEMPLATE = Template('''return {
{% for id, record in records.items() -%}
    [{{ id }}] = { {% for field in fields %}{{ field.name }}={{ record[field.name] }}{% if not loop.last %},{% endif %}{% endfor %} }{% if not loop.last %},{% endif %}
{% endfor -%}
}''')

    def __init__(self, input_dir, output_dir):
        self.input_dir = Path(input_dir)
        self.output_dir = Path(output_dir)
        self._create_dirs()
        # base_names列表, 用于生成UE配置管理器
        self.base_names = []    

    def _create_dirs(self):
        self.data_dir = self.output_dir / "data"
        self.json_dir = self.output_dir / "json"
        self.ue_dir = self.output_dir / "gen_config"
        self.ue_h_dir = self.ue_dir / "public"
        self.ue_cpp_dir = self.ue_dir / "private"
        self.lua_dir = self.output_dir / "gen_lua"
        
        for d in [self.data_dir, self.json_dir, self.ue_dir, self.lua_dir,self.ue_h_dir,self.ue_cpp_dir]:
            d.mkdir(parents=True, exist_ok=True)

    def batch_convert(self):
        """批量转换目录下所有Excel文件"""
        self.base_names = []  # 重置列表
        for excel_file in self.input_dir.glob('*.xlsx'):
            try:
                self.convert_file(excel_file)
                logging.info(f"转换成功: {excel_file.name}")
            except Exception as e:
                logging.error(f"文件转换失败: {excel_file} | {str(e)}")
                sys.exit(1)  # 直接终止程序
        # 生成ConfigMgr
        self.generate_config_mgr()
        # 生成allconfigs.lua
        self.generate_all_configs()
    def convert_file(self, file_path):
        """处理单个文件"""
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 解析数据结构
        base_name = file_path.stem
        fields = self.parse_fields(ws,base_name)
        data = self.parse_data(ws, fields,base_name)
        
         # 生成输出文件
        self.base_names.append(base_name)  # 收集所有base_name
        self.generate_ue_files(base_name, fields)
        self.generate_lua_module(base_name, fields,data)
        self.generate_binary(base_name, fields, data)
        #self.generate_json(base_name, fields, data)

    def parse_fields(self, ws,base_name):
        """解析表头结构"""
        fields = []
        # 获取上下文信息
        filename = base_name
        sheet_name = ws.title
        for col in range(1, ws.max_column+1):
            # 检查导出标记（第三行）
            export_flag = str(ws.cell(row=3, column=col).value).upper()
            if export_flag == 'N':
                continue

            # 获取字段信息
            field = {
                'name': str(ws.cell(row=1, column=col).value).strip(),
                'raw_type': str(ws.cell(row=2, column=col).value).strip(),
                'comment': (ws.cell(row=4, column=col).value or '').strip(),
                'original_col': col  # 记录原始列号
            }
            self._processComment(field) 

            try:
                field['type_info'] = TypeConverter.parse_type(field['raw_type'])
                field['ue_type'] = TypeConverter.get_ue_type(field['type_info'])
                fields.append(field)
            except ValueError as e:
                col_letter = get_column_letter(col)
                logging.error(f"列错误: {base_name} 列 {col_letter} - {str(e)}")
                raise
        # 校验第一列
        if not fields:
            raise ValueError(f"{base_name} 无有效字段")
        self._validate_first_field(fields[0], base_name)
        return fields
    
    def _processComment(self,field):
        comment_lines = field['comment'].split('\n')
        # 格式化每行注释
        formatted_lines = []
        for idx, line in enumerate(comment_lines):
            stripped = line.strip()
            if not stripped:
                continue
                
            # 首行添加注释符
            if idx == 0:
                formatted = f"{stripped}"
            # 后续行对齐注释符
            else:
                formatted = f"  {stripped}"
            
            formatted_lines.append(formatted)
        
        # 合并为统一注释块
        field['comment'] = '\n    '.join(formatted_lines)
        
    def _validate_first_field(self, field, base_name):
        if field['name'].lower() != 'id':
            raise ValueError(f"{base_name} 首列必须为ID")
        if field['type_info'] not in [('basic', 'int32'), ('basic', 'int')]:
            raise ValueError(f"{base_name} ID字段类型必须为int32")
        
    def parse_data(self, ws, fields,base_name):
        """解析数据行（从第五行开始）"""
        data = []
        id_set = set()
        for row_idx in range(5, ws.max_row + 1):
            record = {}
            skip_row = False  # 新增：行跳过标志
            for field in fields:
                raw_value = ws.cell(row=row_idx, column=field['original_col']).value
                 # 特别处理第一列（ID列）
                if field['name'].lower() == 'id':
                    # 新增：空值检查
                    if raw_value is None or str(raw_value).strip() == "":
                        logging.warning(
                            f"跳过空ID行 ➜ 文件 {base_name} 工作表: {ws.title} "
                            f"行: {row_idx}"
                        )
                        skip_row = True
                        break
                record[field['name']] = self._process_value(raw_value, field)
            # 新增：跳过无效行
            if skip_row:
                continue
            self._validate_id(record, fields[0], row_idx, id_set, base_name)
            data.append((row_idx, record))
        return data
    
    def _parse_map(self, value, field):
        parsed = {}
        if value:
            for pair in value.split(','):
                k, v = pair.split(':', 1)
                key_type = field['type_info'][1]
                val_type = field['type_info'][2]
                parsed[TypeConverter.cast_value(k.strip(), key_type)] = TypeConverter.cast_value(v.strip(), val_type)
        return parsed

    def _parse_array(self, value, field):
        return [TypeConverter.cast_value(e.strip(), field['type_info'][1]) 
               for e in value.split(',') if e.strip()] if value else []

    def _validate_id(self, record, id_field, row_idx, id_set, base_name):
        id_value = record[id_field['name']]
        if not isinstance(id_value, int):
            raise ValueError(f"{base_name} 行{row_idx} ID必须为整数")
        if id_value in id_set:
            raise ValueError(f"{base_name} 行{row_idx} ID重复")
        id_set.add(id_value)

    def _process_value(self, raw_value, field):
        cell_value = str(raw_value).strip() if raw_value is not None else ""
        try:
            if field['type_info'][0] == 'map':
                return self._parse_map(cell_value, field)
            elif field['type_info'][0] == 'array':
                return self._parse_array(cell_value, field)
            return TypeConverter.cast_value(cell_value, field['type_info'][1])
        except Exception as e:
            logging.warning(f"值转换警告: {field['name']} - {str(e)}")
            return cell_value
        
    
    def generate_config_mgr(self):
        """生成蓝图函数库"""
        headers = [base for base in self.base_names]
        
        # 生成头文件
        header_content = self.UE_CONFIG_MGR_HEADER.render(
            headers=headers,
            base_names=self.base_names
        )
        (self.ue_h_dir / "ConfigMgr.h").write_text(header_content, encoding='utf-8')

        # 生成实现文件
        cpp_content = self.UE_CONFIG_MGR_CPP.render(
            base_names=self.base_names
        )
        (self.ue_cpp_dir / "ConfigMgr.cpp").write_text(cpp_content, encoding='utf-8')

    def generate_ue_files(self, base_name, fields):
        """生成UE相关文件"""
        struct_name = f"{base_name}Conf"
        manager_name = f"{base_name}ConfManager"
        
        # 生成头文件
        header_content = self.UE_MANAGER_TEMPLATE.render(
            header_file=base_name,
            manager_class=manager_name,
            struct_name=struct_name,
            data_file=base_name,
            fields=fields
        )
        (self.output_dir / f"gen_config/public/{base_name}Conf.h").write_text(header_content, encoding='utf-8')

        # 生成实现文件
        cpp_content = self.UE_MANAGER_CPP_TEMPLATE.render(
            header_file=base_name,
            manager_class=manager_name,
            struct_name=struct_name,
            data_file=base_name,
            fields=fields
        )
        (self.output_dir / f"gen_config/private/{base_name}Conf.cpp").write_text(cpp_content, encoding='utf-8')

    

    def _convert_to_lua_value(self, value, type_info):
        """递归转换值到Lua格式（紧凑模式）"""
        if type_info[0] == 'array':
            return "{" + ",".join(
                self._convert_to_lua_value(item, type_info[1]) 
                for item in value
            ) + "}"
        elif type_info[0] == 'map':
            return "{" + ",".join(
                f"[{self._convert_to_lua_value(k, type_info[1])}]={self._convert_to_lua_value(v, type_info[2])}" 
                for k, v in value.items()
            ) + "}"
        elif isinstance(value, str):
            # 分步处理特殊字符转义
            escaped = value.replace('\\', '\\\\')  # 转义反斜杠
            escaped = escaped.replace('"', '\\"')   # 转义双引号
            escaped = escaped.replace('\n', '\\n')  # 转义换行符
            return f'"{escaped}"'
        elif isinstance(value, bool):
            return "true" if value else "false"
        elif isinstance(value, float):
            return f"{value:.6f}".rstrip('0').rstrip('.') if value % 1 else f"{int(value)}"
        else:
            return str(value)
    
    

    def generate_lua_module(self, base_name, fields, data):
        """生成Lua模块"""
        lua_records = {}
        for excel_row, record in data:
            lua_record = {
                field['name']: self._convert_to_lua_value(record[field['name']], field['type_info'])
                for field in fields
            }
            lua_records[record['id']] = lua_record
        
        content = self.LUA_TEMPLATE.render(records=lua_records, fields=fields)
        content = re.sub(r',\s*}', '}', content)  # 修复最后字段逗号
        (self.output_dir / "gen_lua" / f"{base_name}.lua").write_text(content, encoding='utf-8')

    def generate_all_configs(self):
        content = "return {\n"
        for config in self.base_names:
            content += f'    ["{config}"] = "{config}.lua",\n'
        content += "}"
        (self.lua_dir / "allconfigs.lua").write_text(content, encoding='utf-8')
    
   

    def generate_binary(self, base_name, fields, data):
        """生成二进制文件（带精确错误定位）"""
        buffer = struct.pack('<I', self.MAGIC_NUMBER)
        buffer += struct.pack('<I', self.VERSION)
        buffer += struct.pack('<I', len(data))
        
        # data结构为 (excel_row, record)
        for excel_row, record in data:
            record_buffer = bytearray()
            for field in fields:
                try:
                    # 获取字段值
                    value = record[field['name']]
                    
                    # 执行序列化
                    serialized = BinarySerializer.serialize(value, field['raw_type'])
                    record_buffer += serialized
                    
                except Exception as e:
                    # 转换列号为字母（如3→C）
                    col_letter = get_column_letter(field['original_col'])
                    
                    # 构造详细错误信息
                    error_msg = (
                        f"文件: {base_name}\n"
                        f"工作表: Sheet1\n"  # 假设只处理第一个工作表
                        f"位置: {col_letter}{excel_row}\n"
                        f"字段: {field['name']}({field['raw_type']})\n"
                        f"错误内容: {value}\n"
                        f"错误详情: {str(e)}"
                    )
                    raise ValueError(error_msg)
            
            buffer += struct.pack('<I', len(record_buffer))
            buffer += record_buffer
        
        (self.data_dir / f"{base_name}.data").write_bytes(buffer)

 
    def generate_json(self, base_name, fields, data):
        """生成可观察的JSON数据文件"""
        json_records = []
        
        for excel_row, record_data in data:
            record = {}
            for field in fields:
                field_name = field['name']
                raw_value = record_data[field_name]
                
                try:
                    # 根据字段类型转换数据
                    if field['type_info'][0] == 'basic':
                        converted = TypeConverter.cast_value(raw_value, field['type_info'][1])
                    elif field['type_info'][0] == 'array':
                        converted = [
                            TypeConverter.cast_value(elem, field['type_info'][1]) 
                            for elem in raw_value.split(',') 
                            if elem.strip()
                        ] if raw_value.strip() else []
                    elif field['type_info'][0] == 'map':
                        if not raw_value.strip():  # 处理空值
                            converted = {}
                        else:
                            converted = {}
                            for pair in raw_value.split(','):
                                k, v = pair.split(':', 1)
                                converted[
                                    TypeConverter.cast_value(k.strip(), field['type_info'][1])
                                ] = TypeConverter.cast_value(v.strip(), field['type_info'][2])
                    else:
                        converted = raw_value
                except Exception as e:
                    converted = raw_value
                    logging.warning(
                        f"JSON转换警告 ➜ 文件:{base_name} 行:{excel_row} "
                        f"字段:{field_name} 值:{raw_value} 错误:{str(e)}"
                    )
                
                record[field_name] = converted
            
            json_records.append(record)
    
        json_path = self.json_dir / f"{base_name}_data.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(json_records, f, indent=2, ensure_ascii=False)
        
        logging.info(f"生成观察文件: {json_path}")


excel_path = "D:\\code\\nos3-dev\\GameConfig\\xlsx"        
output_path = "D:\\code\\nos3-dev\\GameConfig\\output"
if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='配置表转换工具')
      # 设置默认值并将 required 设为 False
    parser.add_argument('-i', '--input', default=excel_path, help='输入目录，默认为 ./input')
    parser.add_argument('-o', '--output', default=output_path, help='输出目录，默认为 ./output')
    
    args = parser.parse_args()
    
    converter = ConfigConverter(
        input_dir=args.input,
        output_dir=args.output
    )
    converter.batch_convert()