import argparse
import openpyxl
from pathlib import Path
import logging
import sys

# 配置日志格式
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)

def process_worksheet(ws):
    """处理单个工作表的前两行"""
    for row_num in [1, 2]:  # 处理第1行和第2行
        for cell in ws.iter_cols(min_row=row_num, max_row=row_num):
            if cell[0].value:
                # 保留原始数据类型转换小写
                original = cell[0].value
                if isinstance(original, str):
                    cell[0].value = original.lower()
                elif isinstance(original, (int, float)):
                    cell[0].value = str(original).lower()
                else:
                    cell[0].value = str(original).lower()

def convert_file(file_path):
    """处理单个文件"""
    try:
        # 加载工作簿并保留样式
        wb = openpyxl.load_workbook(file_path, keep_vba=False)
        
        # 处理所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            process_worksheet(ws)
        
        # 覆盖保存原始文件
        wb.save(file_path)
        logging.info(f"成功转换: {file_path}")
        return True
    
    except Exception as e:
        logging.error(f"文件 {file_path} 转换失败: {str(e)}")
        return False

def batch_convert(input_dir):
    """批量转换目录下所有.xlsx文件"""
    input_path = Path(input_dir)
    
    if not input_path.exists():
        logging.error(f"输入目录不存在: {input_dir}")
        return False

    processed_files = 0
    for xlsx_file in input_path.glob('**/*.xlsx'):  # 支持递归遍历
        if convert_file(xlsx_file):
            processed_files += 1

    logging.info(f"转换完成！共处理 {processed_files} 个文件")
    return True

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description='Excel文件批量转换工具 v1.0',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    
    parser.add_argument('-i', '--input',
                        default="D:\\code\\nos3-dev\\nos3-config",
                        help='输入目录路径（支持递归处理子目录）')
    
    args = parser.parse_args()
    
    if not batch_convert(args.input):
        sys.exit(1)