import argparse
import json
import openpyxl
from pathlib import Path
import logging
import sys
from openpyxl.utils import get_column_letter

# 配置日志格式
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)

def is_valid_json_structure(data_str):
    """验证JSON结构是否符合转换要求"""
    try:
        data = json.loads(data_str)
        
        # 1级结构验证
        if not isinstance(data, dict):
            return False
        if 'r_type' not in data or 'itemlist' not in data:
            return False
        
        # 2级结构验证
        itemlist = data.get('itemlist', [])
        if not isinstance(itemlist, list) or len(itemlist) == 0:
            return False
        
        # 3级结构验证
        valid_pairs = []
        for item in itemlist:
            if not isinstance(item, dict):
                continue
            for k, v in item.items():
                # 键必须是数字字符串，值必须是整数
                if str(k).isdigit() and isinstance(v, int):
                    valid_pairs.append(f"{k}:{v}")
        
        return len(valid_pairs) > 0  # 至少包含一个有效键值对
    
    except json.JSONDecodeError:
        return False
    except Exception as e:
        logging.debug(f"JSON验证异常: {str(e)}")
        return False

def convert_column(ws, col_idx):
    """处理单个列的数据转换"""
    # 获取类型定义单元格
    type_cell = ws.cell(row=2, column=col_idx)
    
    # 仅处理string类型列
    if str(type_cell.value).lower() != 'string':
        return False

    modified = False
    processed_rows = 0

    # 从第五行开始处理数据
    for row_idx in range(5, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        raw_value = str(cell.value).strip()
        
        if not raw_value:
            continue

        try:
            # 执行三级验证
            if not is_valid_json_structure(raw_value):
                continue
                
            data = json.loads(raw_value)
            pairs = []
            
            # 提取有效键值对
            for item in data.get('itemlist', []):
                for k, v in item.items():
                    if str(k).isdigit() and isinstance(v, int):
                        pairs.append(f"{k}:{v}")
            
            if pairs:
                # 更新单元格值
                cell.value = ",".join(pairs)
                processed_rows += 1
                modified = True
                
        except Exception as e:
            col_letter = get_column_letter(col_idx)
            logging.warning(f"工作表[{ws.title}] 列{col_letter}行{row_idx} 转换失败: {str(e)}")

    # 更新列类型定义
    if modified:
        type_cell.value = 'map<int32,int32>'
        logging.info(f"工作表[{ws.title}] 列{get_column_letter(col_idx)} 已转换 {processed_rows} 行数据")
    
    return modified

def process_workbook(file_path):
    """处理单个Excel文件"""
    try:
        wb = openpyxl.load_workbook(file_path)
        modified = False
        
        # 遍历所有工作表
        for ws in wb.worksheets:
            # 遍历所有列（从第一列开始）
            for col_idx in range(1, ws.max_column + 1):
                if convert_column(ws, col_idx):
                    modified = True

        # 保存修改
        if modified:
            backup_path = file_path.with_name(f"backup_{file_path.name}")
            file_path.rename(backup_path)  # 创建备份
            wb.save(file_path)
            logging.info(f"文件已转换并备份到: {backup_path}")
        else:
            logging.info(f"无需修改: {file_path}")
            
        return True
    
    except openpyxl.utils.exceptions.InvalidFileException:
        logging.error(f"无效的Excel文件: {file_path}")
    except Exception as e:
        logging.error(f"处理文件异常 {file_path}: {str(e)}")
    
    return False

def batch_convert(input_dir):
    """批量处理目录"""
    input_path = Path(input_dir)
    
    if not input_path.exists():
        raise FileNotFoundError(f"输入目录不存在: {input_dir}")
    
    total_files = 0
    processed_files = 0
    
    # 递归遍历所有.xlsx文件
    for xlsx_file in input_path.rglob('*.xlsx'):
        total_files += 1
        if process_workbook(xlsx_file):
            processed_files += 1

    logging.info(f"处理完成！共找到 {total_files} 个文件，成功转换 {processed_files} 个")
    return processed_files > 0

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description='Excel数据转换工具 v2.0',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    parser.add_argument('-i', '--input', 
                       default="D:\\code\\nos3-dev\\nos3-config",
                       help='输入目录路径（支持递归处理子目录）')
    
    args = parser.parse_args()
    
    try:
        if not batch_convert(args.input):
            sys.exit(1)
    except KeyboardInterrupt:
        logging.warning("用户中断操作")
        sys.exit(1)
    except Exception as e:
        logging.error(f"致命错误: {str(e)}")
        sys.exit(1)